import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from app.models.transaction import Transaction

HEADERS = ["Date", "Posting Date", "Description", "Spent", "Received", "Balance", "Category", "Source", "File"]


def generate_csv(transactions: list[Transaction]) -> io.StringIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for tx in transactions:
        amount = abs(tx.amount)
        spent = f"{amount:.2f}" if tx.type == "debit" else ""
        received = f"{amount:.2f}" if tx.type == "credit" else ""
        source_label = "AI" if tx.category_source == "ai" else "Rule" if tx.category_source == "rule" else "Manual"
        writer.writerow([
            tx.date,
            tx.posting_date or "",
            tx.description,
            spent,
            received,
            f"{tx.balance:.2f}" if tx.balance is not None else "",
            tx.category,
            source_label,
            tx.source or "",
        ])
    output.seek(0)
    return output


def generate_quickbooks_csv(transactions: list[Transaction]) -> io.StringIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Spent", "Received"])
    for tx in transactions:
        amount = abs(tx.amount)
        spent = f"{amount:.2f}" if tx.type == "debit" else ""
        received = f"{amount:.2f}" if tx.type == "credit" else ""
        writer.writerow([tx.date, tx.description, spent, received])
    output.seek(0)
    return output


def generate_excel(transactions: list[Transaction]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    debit_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
    credit_fill = PatternFill(start_color="F2FFF2", end_color="F2FFF2", fill_type="solid")

    for row_idx, tx in enumerate(transactions, 2):
        fill = credit_fill if tx.type == "credit" else debit_fill
        amount = abs(tx.amount)

        ws.cell(row=row_idx, column=1, value=tx.date).fill = fill
        ws.cell(row=row_idx, column=2, value=tx.posting_date or "").fill = fill
        ws.cell(row=row_idx, column=3, value=tx.description).fill = fill

        # Spent column (debit)
        spent_cell = ws.cell(row=row_idx, column=4, value=amount if tx.type == "debit" else None)
        spent_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        spent_cell.fill = fill

        # Received column (credit)
        received_cell = ws.cell(row=row_idx, column=5, value=amount if tx.type == "credit" else None)
        received_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        received_cell.fill = fill

        if tx.balance is not None:
            balance_cell = ws.cell(row=row_idx, column=6, value=tx.balance)
            balance_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            balance_cell.fill = fill
        else:
            ws.cell(row=row_idx, column=6, value="").fill = fill

        ws.cell(row=row_idx, column=7, value=tx.category).fill = fill

        source_label = "AI" if tx.category_source == "ai" else "Rule" if tx.category_source == "rule" else "Manual"
        ws.cell(row=row_idx, column=8, value=source_label).fill = fill
        ws.cell(row=row_idx, column=9, value=tx.source or "").fill = fill

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
